import os
import json
import re
import uuid
from email import policy
from email.parser import BytesParser

from flask import Flask, render_template, request, redirect, url_for, send_file, flash, jsonify
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dateutil import parser as date_parser
from dotenv import load_dotenv
import google.generativeai as genai

# Load .env
load_dotenv()

app = Flask(__name__)
app.secret_key = "hermes-secret-key-2026"

UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads")
OUTPUT_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
DATA_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
SESSIONS_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sessions")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
os.makedirs(SESSIONS_FOLDER, exist_ok=True)

# ── Configure Gemini ──
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "")
if GEMINI_API_KEY:
    genai.configure(api_key=GEMINI_API_KEY)


# ──────────────────────────────────────────────
# Sender-type classification keywords
# ──────────────────────────────────────────────
PROFESSOR_KEYWORDS = [
    "prof", "professor", "dr.", "doctor", "dean", "hod",
    "faculty", "lecturer", "department", "dept",
    "university", "institute", "college", "academic",
    "pvc.", "vc.", "chancellor", "principal",
    ".edu", ".ac.", "cumail.in", "culko.in",
]

STUDENT_KEYWORDS = [
    "student", "b.tech", "btech", "m.tech", "mtech",
    "bsc", "msc", "phd", "scholar", "roll no",
    "enrollment", "batch", "semester", "classmate",
]

COMPANY_KEYWORDS = [
    "internshala", "linkedin", "naukri", "indeed",
    "glassdoor", "hiring", "recruiter", "hr@",
    "careers", "talent", "opportunity", "job",
    "amazon", "google", "microsoft", "github",
    "schbang", "noreply", "no-reply", "newsletter",
    "marketing", "team@", "support@", "info@",
    "notifications", "digest", "updates",
    "hackclub", "hack club",
]

# ──────────────────────────────────────────────
# Purpose classification keywords
# ──────────────────────────────────────────────
PURPOSE_MAP = {
    "Leave / Permission": [
        "leave", "permission", "absent", "absence",
        "sick", "medical", "casual leave", "duty leave",
        "on leave", "request for leave", "approval",
        "drone fly", "fly permission",
    ],
    "Event / Competition": [
        "event", "competition", "hackathon", "fest",
        "techkriti", "workshop", "seminar", "webinar",
        "conference", "meetup", "contest", "challenge",
        "quiz", "coding", "drone racing", "multirotor",
        "techfest", "code", "qualifier",
    ],
    "Job / Internship": [
        "job", "internship", "intern", "hiring",
        "opportunity", "career", "vacancy", "apply",
        "placement", "openings", "recruiter", "resume",
        "ctc", "salary", "compensation", "position",
    ],
    "Account / Service": [
        "account", "verify", "confirm", "password",
        "login", "sign in", "sign up", "register",
        "welcome", "activation", "subscription",
        "onshape", "github education", "archive",
        "data request", "login code",
    ],
    "Payment / Finance": [
        "payment", "fee", "invoice", "receipt",
        "transaction", "amount", "pay", "dues",
        "refund", "billing",
    ],
    "Social / Messaging": [
        "messaged you", "sent you", "message from",
        "chat", "dm", "direct message", "replied",
        "mentioned you", "tagged you",
    ],
    "Newsletter / Promotional": [
        "newsletter", "digest", "update", "weekly",
        "monthly", "announcement", "promo", "offer",
        "discount", "deal", "sale", "unsubscribe",
        "arcade", "skills",
    ],
}


def extract_text_from_eml(file_path):
    """Parse an .eml file and return structured data."""
    with open(file_path, "rb") as f:
        msg = BytesParser(policy=policy.default).parse(f)

    subject = msg.get("subject", "(No Subject)")
    from_raw = msg.get("from", "")
    to_raw = msg.get("to", "")
    cc_raw = msg.get("cc", "")
    date_raw = msg.get("date", "")

    # Extract plain-text body
    body = ""
    if msg.is_multipart():
        for part in msg.walk():
            ctype = part.get_content_type()
            if ctype == "text/plain":
                payload = part.get_payload(decode=True)
                if payload:
                    body = payload.decode("utf-8", errors="replace")
                    break
        if not body:
            for part in msg.walk():
                ctype = part.get_content_type()
                if ctype == "text/html":
                    payload = part.get_payload(decode=True)
                    if payload:
                        html_text = payload.decode("utf-8", errors="replace")
                        body = re.sub(r"<[^>]+>", " ", html_text)
                        body = re.sub(r"\s+", " ", body).strip()
                        break
    else:
        payload = msg.get_payload(decode=True)
        if payload:
            body = payload.decode("utf-8", errors="replace")

    # Parse date
    parsed_date = ""
    if date_raw:
        try:
            dt = date_parser.parse(date_raw)
            parsed_date = dt.strftime("%Y-%m-%d %H:%M")
        except Exception:
            parsed_date = date_raw

    return {
        "subject": subject,
        "from": from_raw,
        "to": to_raw,
        "cc": cc_raw,
        "date": parsed_date,
        "body": body[:3000],
    }


def classify_sender(email_data):
    """Classify the sender as Professor, Student, Company, or Other."""
    haystack = (
        email_data["from"].lower()
        + " " + email_data["body"][:800].lower()
        + " " + email_data["subject"].lower()
    )
    scores = {"Professor": 0, "Student": 0, "Company": 0}
    for kw in PROFESSOR_KEYWORDS:
        if kw in haystack:
            scores["Professor"] += 1
    for kw in STUDENT_KEYWORDS:
        if kw in haystack:
            scores["Student"] += 1
    for kw in COMPANY_KEYWORDS:
        if kw in haystack:
            scores["Company"] += 1

    max_score = max(scores.values())
    if max_score == 0:
        return "Other"
    for cat in ["Professor", "Company", "Student"]:
        if scores[cat] == max_score:
            return cat
    return "Other"


def classify_purpose(email_data):
    """Classify the email purpose."""
    haystack = (
        email_data["subject"].lower()
        + " " + email_data["body"][:1500].lower()
    )
    scores = {}
    for purpose, keywords in PURPOSE_MAP.items():
        score = sum(1 for kw in keywords if kw in haystack)
        scores[purpose] = score

    max_score = max(scores.values())
    if max_score == 0:
        return "General / Other"
    return max(scores, key=scores.get)


def process_eml_files(folder_path):
    """Process all .eml files in a folder and return classified data."""
    results = []
    eml_files = [f for f in os.listdir(folder_path) if f.lower().endswith(".eml")]

    for filename in sorted(eml_files):
        filepath = os.path.join(folder_path, filename)
        try:
            data = extract_text_from_eml(filepath)
            sender_type = classify_sender(data)
            purpose = classify_purpose(data)
            results.append({
                "filename": filename,
                "subject": data["subject"],
                "from": data["from"],
                "to": data["to"],
                "date": data["date"],
                "sender_type": sender_type,
                "purpose": purpose,
                "body_preview": data["body"][:200],
                "body_full": data["body"],
            })
        except Exception as e:
            results.append({
                "filename": filename,
                "subject": f"[Error: {str(e)}]",
                "from": "", "to": "", "date": "",
                "sender_type": "Unknown",
                "purpose": "Unknown",
                "body_preview": "",
                "body_full": "",
            })
    return results


def save_session(session_id, results):
    """Save processed results to a JSON file so we can retrieve them at export."""
    path = os.path.join(SESSIONS_FOLDER, f"{session_id}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False)


def load_session(session_id):
    """Load processed results from a session file."""
    path = os.path.join(SESSIONS_FOLDER, f"{session_id}.json")
    if not os.path.isfile(path):
        return None
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


# ──────────────────────────────────────────────
# Gemini Flash — Generate email response
# ──────────────────────────────────────────────
def generate_ai_response(email_data):
    """Use Gemini Flash to draft a reply for an email."""
    if not GEMINI_API_KEY:
        return "[No API key configured]"

    prompt = f"""You are a helpful email assistant. Draft a professional and appropriate reply 
to the following email. Keep the tone matching the context — formal for professors/officials, 
friendly for students, professional for companies. Be concise but complete.

--- EMAIL ---
From: {email_data.get('from', 'Unknown')}
Subject: {email_data.get('subject', '(No Subject)')}
Date: {email_data.get('date', '')}
Category: {email_data.get('sender_type', '')} — {email_data.get('purpose', '')}

Body:
{email_data.get('body_full', email_data.get('body_preview', ''))}
--- END EMAIL ---

Draft a reply (just the reply body, no subject line):"""

    try:
        model = genai.GenerativeModel("gemini-2.0-flash")
        response = model.generate_content(prompt)
        return response.text.strip()
    except Exception as e:
        return f"[AI Error: {str(e)}]"


# ──────────────────────────────────────────────
# Excel Generation
# ──────────────────────────────────────────────
def generate_excel(results, output_path, ai_responses=None):
    """Generate a beautifully formatted Excel workbook.
    ai_responses: dict mapping index (int) -> response string
    """
    wb = Workbook()

    HEADER_FILL = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    HEADER_FONT = Font(name="Segoe UI", bold=True, color="FFFFFF", size=12)
    EVEN_ROW = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")
    ODD_ROW = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    AI_FILL = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    SENDER_COLORS = {
        "Professor": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
        "Student":   PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
        "Company":   PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
        "Other":     PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),
    }
    SENDER_FONTS = {
        "Professor": Font(name="Segoe UI", bold=True, color="2E7D32", size=11),
        "Student":   Font(name="Segoe UI", bold=True, color="1565C0", size=11),
        "Company":   Font(name="Segoe UI", bold=True, color="E65100", size=11),
        "Other":     Font(name="Segoe UI", bold=True, color="6A1B9A", size=11),
    }

    has_ai = ai_responses and len(ai_responses) > 0

    # ── Sheet: All Emails ──
    ws_all = wb.active
    ws_all.title = "All Emails"
    headers = ["#", "Subject", "From", "To", "Date", "Sender Type", "Purpose", "Body Preview"]
    col_widths = [5, 50, 35, 35, 18, 15, 22, 60]

    if has_ai:
        headers.append("AI Generated Response")
        col_widths.append(80)

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws_all.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws_all.column_dimensions[cell.column_letter].width = width

    ws_all.row_dimensions[1].height = 30

    for row_idx, item in enumerate(results, 2):
        i = row_idx - 2  # 0-based index
        row_data = [
            row_idx - 1,
            item["subject"],
            item["from"],
            item["to"],
            item["date"],
            item["sender_type"],
            item["purpose"],
            item.get("body_preview", ""),
        ]

        if has_ai:
            response_text = ai_responses.get(str(i), ai_responses.get(i, ""))
            row_data.append(response_text)

        fill = EVEN_ROW if row_idx % 2 == 0 else ODD_ROW
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_all.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Segoe UI", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
            cell.fill = fill

        # Colour the sender-type cell
        sender_cell = ws_all.cell(row=row_idx, column=6)
        st = item["sender_type"]
        sender_cell.fill = SENDER_COLORS.get(st, ODD_ROW)
        sender_cell.font = SENDER_FONTS.get(st, Font(name="Segoe UI", size=10))
        sender_cell.alignment = Alignment(horizontal="center", vertical="top")

        # Highlight AI response cell
        if has_ai and row_data[-1]:
            ai_cell = ws_all.cell(row=row_idx, column=len(row_data))
            ai_cell.fill = AI_FILL
            ai_cell.font = Font(name="Segoe UI", size=10, italic=True, color="5D4037")

    ws_all.auto_filter.ref = f"A1:{chr(64+len(headers))}{len(results)+1}"

    # ── Sheets by Sender Type ──
    for sender_type in ["Professor", "Student", "Company", "Other"]:
        filtered = [(i, r) for i, r in enumerate(results) if r["sender_type"] == sender_type]
        if not filtered:
            continue
        ws = wb.create_sheet(title=f"{sender_type}s")
        sh = ["#", "Subject", "From", "Date", "Purpose", "Body Preview"]
        sw = [5, 50, 35, 18, 22, 60]
        if has_ai:
            sh.append("AI Generated Response")
            sw.append(80)

        for col_idx, (header, width) in enumerate(zip(sh, sw), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
            ws.column_dimensions[cell.column_letter].width = width
        ws.row_dimensions[1].height = 30

        for row_idx, (orig_idx, item) in enumerate(filtered, 2):
            row_data = [row_idx - 1, item["subject"], item["from"], item["date"], item["purpose"], item.get("body_preview", "")]
            if has_ai:
                row_data.append(ai_responses.get(str(orig_idx), ai_responses.get(orig_idx, "")))

            fill = EVEN_ROW if row_idx % 2 == 0 else ODD_ROW
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name="Segoe UI", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = THIN_BORDER
                cell.fill = fill

            if has_ai and row_data[-1]:
                ai_cell = ws.cell(row=row_idx, column=len(row_data))
                ai_cell.fill = AI_FILL
                ai_cell.font = Font(name="Segoe UI", size=10, italic=True, color="5D4037")

    # ── Sheets by Purpose ──
    all_purposes = sorted(set(r["purpose"] for r in results))
    for purpose in all_purposes:
        filtered = [(i, r) for i, r in enumerate(results) if r["purpose"] == purpose]
        if not filtered:
            continue
        safe_title = purpose.replace("/", "-").replace("\\", "-")[:31]
        ws = wb.create_sheet(title=safe_title)
        sh = ["#", "Subject", "From", "Date", "Sender Type", "Body Preview"]
        sw = [5, 50, 35, 18, 15, 60]
        if has_ai:
            sh.append("AI Generated Response")
            sw.append(80)

        for col_idx, (header, width) in enumerate(zip(sh, sw), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
            ws.column_dimensions[cell.column_letter].width = width
        ws.row_dimensions[1].height = 30

        for row_idx, (orig_idx, item) in enumerate(filtered, 2):
            row_data = [row_idx - 1, item["subject"], item["from"], item["date"], item["sender_type"], item.get("body_preview", "")]
            if has_ai:
                row_data.append(ai_responses.get(str(orig_idx), ai_responses.get(orig_idx, "")))

            fill = EVEN_ROW if row_idx % 2 == 0 else ODD_ROW
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name="Segoe UI", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = THIN_BORDER
                cell.fill = fill

            if has_ai and row_data[-1]:
                ai_cell = ws.cell(row=row_idx, column=len(row_data))
                ai_cell.fill = AI_FILL
                ai_cell.font = Font(name="Segoe UI", size=10, italic=True, color="5D4037")

    wb.save(output_path)
    return output_path


# ─────────────────────────────────────────────
# Flask Routes
# ─────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/upload", methods=["POST"])
def upload():
    """Handle .eml file uploads and process them."""
    files = request.files.getlist("eml_files")

    if not files or files[0].filename == "":
        flash("Please select at least one .eml file.", "error")
        return redirect(url_for("index"))

    session_id = uuid.uuid4().hex[:12]
    session_folder = os.path.join(UPLOAD_FOLDER, session_id)
    os.makedirs(session_folder, exist_ok=True)

    saved_count = 0
    for f in files:
        if f.filename and f.filename.lower().endswith(".eml"):
            safe_name = f.filename.replace("/", "_").replace("\\", "_")
            f.save(os.path.join(session_folder, safe_name))
            saved_count += 1

    if saved_count == 0:
        flash("No valid .eml files found in your upload.", "error")
        return redirect(url_for("index"))

    results = process_eml_files(session_folder)
    save_session(session_id, results)

    return render_template(
        "results.html",
        results=results,
        total=len(results),
        session_id=session_id,
    )


@app.route("/process-data", methods=["POST"])
def process_data():
    """Process the .eml files already in the data/ folder."""
    if not os.path.isdir(DATA_FOLDER):
        flash("No data/ folder found.", "error")
        return redirect(url_for("index"))

    eml_files = [f for f in os.listdir(DATA_FOLDER) if f.lower().endswith(".eml")]
    if not eml_files:
        flash("No .eml files found in the data/ folder.", "error")
        return redirect(url_for("index"))

    results = process_eml_files(DATA_FOLDER)

    session_id = "data_" + uuid.uuid4().hex[:8]
    save_session(session_id, results)

    return render_template(
        "results.html",
        results=results,
        total=len(results),
        session_id=session_id,
    )


@app.route("/export", methods=["POST"])
def export():
    """Export selected emails to Excel, generating AI responses for checked ones."""
    session_id = request.form.get("session_id", "")
    selected_raw = request.form.getlist("selected_indices")  # list of index strings

    results = load_session(session_id)
    if not results:
        flash("Session expired or not found. Please process emails again.", "error")
        return redirect(url_for("index"))

    selected_indices = set()
    for s in selected_raw:
        try:
            selected_indices.add(int(s))
        except ValueError:
            pass

    # Generate AI responses for selected emails
    ai_responses = {}
    for idx in selected_indices:
        if 0 <= idx < len(results):
            ai_responses[idx] = generate_ai_response(results[idx])

    # Generate Excel
    output_filename = f"Hermes_Report_{session_id}.xlsx"
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)
    generate_excel(results, output_path, ai_responses)

    return send_file(output_path, as_attachment=True, download_name=output_filename)


@app.route("/download/<filename>")
def download(filename):
    """Download a generated Excel file."""
    filepath = os.path.join(OUTPUT_FOLDER, filename)
    if not os.path.isfile(filepath):
        flash("File not found.", "error")
        return redirect(url_for("index"))
    return send_file(filepath, as_attachment=True, download_name=filename)


if __name__ == "__main__":
    app.run(debug=True, port=5000)
